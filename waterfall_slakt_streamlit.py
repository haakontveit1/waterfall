import pandas as pd
from datetime import datetime, timedelta, date
import numpy as np
import matplotlib.pyplot as plt
import streamlit as st
from openpyxl import load_workbook

def les_data(uploaded_file):
    if uploaded_file is not None:
        try:
            if sheet_type == "slakt":
                # Load data and extract comments for "slakt"
                df = pd.read_excel(uploaded_file, header=2)
                workbook = load_workbook(uploaded_file)
                sheet = workbook.active

                comments = []
                for idx, row in enumerate(sheet.iter_rows(min_row=3, min_col=4, max_col=4), start=0):
                    cell = row[0]
                    if cell.comment:
                        comment_text = cell.comment.text
                        colon_count = 0
                        hh_mm = ""
                        for char in comment_text:
                            if char == ":":
                                colon_count += 1
                            if colon_count >= 3 and char.isdigit():
                                hh_mm += char
                        comment_text = hh_mm.strip() if hh_mm else ""
                    else:
                        comment_text = ""
                    comments.append(comment_text)

                aligned_comments = comments[1:] + [""]
                df["comments"] = aligned_comments[:len(df)]
            else:
                # Simple load for "filet"
                df = pd.read_excel(uploaded_file, header=2)

            return df
        except Exception as e:
            st.error(f"Feil ved lesing av Excel-filen: {e}")
            return None
    return None


def velg_dato():
    år = st.number_input("Velg år:", min_value=2024, max_value=datetime.now().year)
    maneder = ["Januar", "Februar", "Mars", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Desember"]
    maaned = st.selectbox("Velg måned:", list(range(1, 13)), format_func=lambda x: maneder[x-1])
    dag = st.number_input("Velg dag (1-31):", min_value=1, max_value=31)
    valgt_dato = datetime(år, maaned, dag)
    return valgt_dato

def hent_uke_dager(år, uke_nummer):
    dager = []

    # Handle edge case for week 1
    if uke_nummer == 1:
        forrige_år = år - 1
        uke_forrige_år = 52
        if date(forrige_år, 12, 28).isocalendar()[1] == 53:
            uke_forrige_år = 53
        # Find Thursday of the last week of the previous year using isocalendar
        torsdag_forrige_uke = date.fromisocalendar(forrige_år, uke_forrige_år, 4)
    else:
        # Normal case: Find Thursday of the previous week using isocalendar
        torsdag_forrige_uke = date.fromisocalendar(år, uke_nummer - 1, 4)

    # Convert to datetime and add Thursday and Friday from the previous week
    torsdag_forrige_uke_dt = datetime.combine(torsdag_forrige_uke, datetime.min.time())
    dager += [torsdag_forrige_uke_dt, torsdag_forrige_uke_dt + timedelta(days=1)]  # Thursday and Friday

    # Check and add Saturday and Sunday from the previous week if they exist in the production data
    lørdag_forrige_uke = torsdag_forrige_uke_dt + timedelta(days=2)  # Saturday
    søndag_forrige_uke = torsdag_forrige_uke_dt + timedelta(days=3)  # Sunday
    if lørdag_forrige_uke.date() in df['Dato'].values:
        dager.append(lørdag_forrige_uke)
    if søndag_forrige_uke.date() in df['Dato'].values:
        dager.append(søndag_forrige_uke)

    # Add Monday, Tuesday, and Wednesday from the current week using isocalendar
    try:
        # Calculate the Monday of the given ISO week directly
        monday = date.fromisocalendar(år, uke_nummer, 1)
        dager += [datetime.combine(monday + timedelta(days=i), datetime.min.time()) for i in range(3)]
    except ValueError as e:
        st.write(f"Error calculating days for year {år}, week {uke_nummer}: {e}")
    date_only = []
    for i in dager:
        date_only.append(i.strftime("%Y-%m-%d"))
    return date_only



# def hent_uke_dager(år, uke_nummer):
#     try:
#         # Find the Monday of the given ISO week
#         mandag = date.fromisocalendar(år, uke_nummer, 1)
        
#         # Append all days from Monday to Sunday
#         dager = [mandag + timedelta(days=i) for i in range(7)]
#         date_only = []
#         for i in dager:
#             date_only.append(i.strftime("%Y-%m-%d"))
#         return date_only
#     except ValueError as e:
#         print(f"Feil ved henting av dager for uken: {e}")
#         return []


def hent_maned_dager(år, maned):
    dager = []
    try:
        for _, row in df.iterrows():
            dato = row['Dato']
            if dato.year == år and dato.month == maned:
                dager.append(dato)
    except Exception as e:
        st.error(f"Feil ved henting av dager for måneden: {e}")
    return dager


def beregn_stopptid(row):
    try:
        if sheet_type == "slakt":
            stopptid = (
                row.iloc[27:31].fillna(0).sum() +
                row.iloc[34:40].fillna(0).sum() / 6 +
                row.iloc[40:51].fillna(0).sum()
            )
        elif sheet_type == "filet":
            stopptid = row.iloc[32:52].fillna(0).sum()
        return stopptid
    except Exception as e:
        st.error(f"Feil ved beregning av stopptid: {e}")
        return None
    

def beregn_faktiskproduksjon(row):
    try:
        if sheet_type == "slakt":
            # Handle edge cases for "slakt"
            start_time = datetime.strptime(str(row.iloc[2]), "%H:%M:%S")
            end_time_cell_value = str(row.iloc[3])

            if end_time_cell_value in ["23:59:00", "00:00:00"]:
                comment_text = row['comments']

                if comment_text:
                    try:
                        if len(comment_text) == 4 and comment_text.isdigit():
                            hh = int(comment_text[:2])
                            mm = int(comment_text[2:])
                            parsed_time = timedelta(hours=hh, minutes=mm)
                        elif ":" in comment_text:
                            hh, mm = map(int, comment_text.split(":"))
                            parsed_time = timedelta(hours=hh, minutes=mm)
                        else:
                            raise ValueError("Invalid time format in comment.")
                    except Exception as e:
                        st.error(f"Could not parse time from comment: {e}")
                        return None, None

                    if end_time_cell_value == "23:59:00":
                        end_time = timedelta(minutes=0) + parsed_time
                    else:  # "00:00:00"
                        end_time = timedelta(hours=24) + parsed_time

                    end_time_datetime = datetime.combine(start_time.date(), datetime.min.time()) + end_time
                    if end_time_datetime < start_time:
                        end_time_datetime += timedelta(days=1)

                    work_duration = end_time_datetime - start_time
                else:
                    st.error("Sluttidspunkt skrevet kan indikere sluttid etter kl 00:00, men ingen kommentar funnet.")
                    end_time_datetime = datetime.strptime(end_time_cell_value, "%H:%M:%S")
                    work_duration = end_time_datetime - start_time
            else:
                try:
                    # Attempt to parse with the default format
                    end_time_datetime = datetime.strptime(end_time_cell_value, "%H:%M:%S")
                except ValueError:
                    # Handle alternative time formats, including the specific '1900-01-01 00:00:00'
                    possible_formats = ["%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S"]
                    for fmt in possible_formats:
                        try:
                            end_time_datetime = datetime.strptime(end_time_cell_value, fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        st.error(f"Kunne ikke parse sluttidspunkt: {end_time_cell_value}")
                        return None, None

                work_duration = end_time_datetime - start_time
                
            arbeidstimer = work_duration.total_seconds() / 60
            antall_fisk = row.iloc[5]

        elif sheet_type == "filet":
            start_time = datetime.strptime(str(row.iloc[6]), "%H:%M:%S")
            end_time_cell_value = str(row.iloc[7])

            try:
                end_time_datetime = datetime.strptime(end_time_cell_value, "%H:%M:%S")
            except ValueError:
                possible_formats = ["%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S"]
                for fmt in possible_formats:
                    try:
                        end_time_datetime = datetime.strptime(end_time_cell_value, fmt)
                        break
                    except ValueError:
                        continue
                else:
                    st.error(f"Kunne ikke parse sluttidspunkt: {end_time_cell_value}")
                    return None, None

            work_duration = end_time_datetime - start_time
            arbeidstimer = work_duration.total_seconds() / 60
            antall_fisk = row.iloc[12]
        if arbeidstimer < 0:
            arbeidstimer += 24*60 
        
        return arbeidstimer, antall_fisk

    except Exception as e:
        st.error(f"Feil ved beregning av faktisk produksjon: {e}")
        return None, None

    
def pen_dato(date):
    # Define Norwegian month names
    months = {
        "January": "Januar", "February": "Februar", "March": "Mars", "April": "April",
        "May": "Mai", "June": "Juni", "July": "Juli", "August": "August",
        "September": "September", "October": "Oktober", "November": "November", "December": "Desember"
    }

    # Get the English day and month names
    day_name = date.strftime("%A")
    month_name = date.strftime("%B")

    # Explicitly handle known encoding issues
    days = {
        "Monday": "Mandag", "Tuesday": "Tirsdag", "Wednesday": "Onsdag",
        "Thursday": "Torsdag", "Friday": "Fredag", "Saturday" : "Lørdag", "Sunday" : "Søndag"
        }
    norwegian_day = days.get(day_name, day_name)

    # Map month name
    norwegian_month = months.get(month_name, month_name)

    # Format the date
    formatted_date = f"{norwegian_day} {date.day}. {norwegian_month} {date.year}"

    # Final check for encoding issues
    # Replace incorrectly decoded characters, just in case
    formatted_date = formatted_date.replace("sÃ¸ndag", "Søndag").replace("lÃ¸rdag", "Lørdag")

    return formatted_date

    
def lag_graph(annet,faktisk_takt, stopptid_takt, dag, graf_type):

    if sheet_type == "slakt":
        tittel = "på slakt"
        fisk = "fisk"
        
    
    elif sheet_type == "filet":
        tittel = "på filet"
        fisk = "filet"
    
    
    fig, ax = plt.subplots(figsize=(10, 5), dpi=100)
    stages = ['100% OEE', 'Stopptid', 'Annet']
    values = [oee_100, -stopptid_takt, -annet]
    cum_values = np.cumsum([0] + values).tolist()
    value_starts = cum_values[:-1]
    colors = ['blue', 'red', 'orange']
    
    
    for i in range(len(stages)):
        ax.bar(stages[i], values[i], bottom=value_starts[i], color=colors[i], edgecolor='black')

    ax.bar('Takttid', faktisk_takt, bottom=0, color='green', edgecolor='black')
    ax.bar('Takttid', stiplet_hoyde - faktisk_takt, bottom=faktisk_takt, color='none', edgecolor='green', hatch='//')

    for i in range(len(stages)):
        if stages[i] == 'Stopptid':
            if sheet_type == "slakt":
                if stopptid_takt < 9:
                    # Place the text outside the bar if the value is less than 7
                    dynamic_offset = value_starts[i] + values[i] -2  # Adjust `-10` for spacing
                    ax.text(
                        stages[i], dynamic_offset,
                        f'{values[i]} ({abs(values[i]) / oee_100 * 100:.1f}%)',
                        ha='center', va='top', color='black', fontweight='bold'
                        )
                else:
                    # Place the text inside the bar if the value is greater than or equal to 7
                    y_pos = value_starts[i] + values[i] / 2
                    ax.text(
                        stages[i], y_pos,
                        f'{values[i]} ({abs(values[i]) / oee_100 * 100:.1f}%)',
                        ha='center', va='center', color='black', fontweight='bold'
                    )
            elif sheet_type == "filet":
                if stopptid_takt < 1.5:
                    # Place the text outside the bar if the value is less than 7
                        dynamic_offset = value_starts[i] + values[i]-0.5 # Adjust `-10` for spacing
                        ax.text(
                            stages[i], dynamic_offset,
                            f'{values[i]} ({abs(values[i]) / oee_100 * 100:.1f}%)',
                            ha='center', va='top', color='black', fontweight='bold'
                        )
                else:
                    # Place the text inside the bar if the value is greater than or equal to 7
                    y_pos = value_starts[i] + values[i] / 2
                    ax.text(
                        stages[i], y_pos,
                        f'{values[i]} ({abs(values[i]) / oee_100 * 100:.1f}%)',
                        ha='center', va='center', color='black', fontweight='bold'
                    )
        else:
            # Place the text inside other bars as before
            y_pos = value_starts[i] + values[i] / 2
            ax.text(
                stages[i], y_pos,
                f'{values[i]} ({abs(values[i]) / oee_100 * 100:.1f}%)',
                ha='center', va='center', color='black', fontweight='bold'
            )

    ax.text(
        'Takttid', faktisk_takt / 2,
        f'{faktisk_takt} ({(faktisk_takt / oee_100 * 100):.1f}%)',
        ha='center', va='center', color='black', fontweight='bold'
    )

    gap_to_80 = stiplet_hoyde - faktisk_takt
    ax.text(
        'Takttid', stiplet_hoyde + 1,
        f'{round(gap_to_80, 2)} ({(gap_to_80 / oee_100 * 100):.1f}%)',
        ha='center', va='bottom', color='green', fontweight='bold'
    )
    
    
    if graf_type == "enkeltgraf":
        dag = pen_dato(dag)
        ax.set_ylabel(f'Antall {fisk} produsert per minutt')
        ax.set_title(f'Daglig produksjon {tittel} {dag}')
        st.pyplot(fig)
        
    elif graf_type == "ukesnitt":
        ax.set_ylabel(f'Antall {fisk} produsert per minutt')
        ax.set_title(f'Ukentlig gjennomsnitt {tittel} for uke {dag}')
        st.pyplot(fig)
        
    elif graf_type == "manedsnitt":
        ax.set_ylabel(f'Antall {fisk} produsert per minutt')
        ax.set_title(f'Månedlig gjennomsnitt {tittel}  i {dag}')
        st.pyplot(fig)
        
def enkelt_dato():
    valgt_dato = velg_dato()
    valgt_dato_enkel = valgt_dato.date()
    if valgt_dato_enkel in df['Dato'].values:
        row = df[df['Dato'] == valgt_dato_enkel].iloc[0]
        stopptid = beregn_stopptid(row)
        arbeidstimer, antall_fisk = beregn_faktiskproduksjon(row)
        if stopptid is None or arbeidstimer is None or antall_fisk is None:
            st.error("Kan ikke beregne verdier. Sjekk om du har valgt riktig filtype og lastet opp riktig fil.")
            return
        
        stopptid_impact = stopptid * oee_100
        stopptid_takt = round(stopptid_impact / arbeidstimer, 2)
        faktisk_takt = round(antall_fisk / arbeidstimer, 2)
        kjente_faktorer = round(stopptid_takt, 2)
        annet = oee_100 - kjente_faktorer - faktisk_takt
        annet = round(annet, 2)
        graf_type = "enkeltgraf"
        
        lag_graph(
            annet,
            faktisk_takt, stopptid_takt,
            valgt_dato,graf_type)
    else:
        st.warning("Datoen du valgte finnes ikke i input-arket. Dette er enten fordi du tastet inn en ugyldig dato eller fordi datoen ikke hadde noen produksjon (eks helg).")
    return        
    

def uke():
    year = st.number_input("Velg år:", min_value=2024, max_value=datetime.now().year)
    week_number = st.number_input("Velg uke nummer:", min_value=1, max_value=52)
    week_days = hent_uke_dager(year, week_number)
    daglig_data = []
    for dag in week_days:
        dag = datetime.strptime(dag, "%Y-%m-%d").date()
        if dag in df['Dato'].values:
            row = df[df['Dato'] == dag].iloc[0]
            # Format with month as text
            formatted_date = dag.strftime("%d. %B %Y")
            # Print in Norwegian style
            st.write(f"Dato: {formatted_date}")
            stopptid = beregn_stopptid(row)
            arbeidstimer, antall_fisk = beregn_faktiskproduksjon(row)
            if stopptid is None or arbeidstimer is None or antall_fisk is None:
                st.error(f"Kan ikke beregne verdier for {dag.strftime('%d.%m.%Y')}. Sjekk om du har valgt riktig filtype og lastet opp riktig fil.")
                return
            daglig_data.append((dag, stopptid, arbeidstimer, antall_fisk))
            
            stopptid_impact = stopptid * oee_100
            stopptid_takt = round(stopptid_impact / arbeidstimer, 2)
            faktisk_takt = round(antall_fisk / arbeidstimer, 2)
            kjente_faktorer = round(stopptid_takt, 2)
            annet = oee_100 - kjente_faktorer - faktisk_takt
            annet = round(annet, 2)
            graf_type = "enkeltgraf"
            lag_graph(
                annet,
                faktisk_takt, stopptid_takt,
                dag, graf_type)

    if not daglig_data:
        st.warning("Ingen gyldige data funnet for den valgte uken.")
        return
    

    # Weekly
    avg_stopptid = np.mean([data[1] for data in daglig_data])
    avg_arbeidstimer = np.mean([data[2] for data in daglig_data])
    avg_antall_fisk = np.mean([data[3] for data in daglig_data])

    avg_stopptid_impact = avg_stopptid * oee_100
    avg_stopptid_takt = round(avg_stopptid_impact / avg_arbeidstimer, 2)
    avg_faktisk_takt = round(avg_antall_fisk / avg_arbeidstimer, 2)
    avg_kjente_faktorer = round(avg_stopptid_takt, 2)
    avg_annet = oee_100 - avg_kjente_faktorer - avg_faktisk_takt
    avg_annet = round(avg_annet, 2)
    graf_type = "ukesnitt"
    
    lag_graph(
        avg_annet,
        avg_faktisk_takt, avg_stopptid_takt,
        week_number, graf_type
        )
    
    
def maned(): 
    year = st.number_input("Velg år:", min_value=2024, max_value=datetime.now().year)
    months = ["Velg måned", "Januar", "Februar", "Mars", "April", "Mai", "Juni", 
              "Juli", "August", "September", "Oktober", "November", "Desember"]
    selected_month = st.selectbox("Velg måned:", months)
    
    # Check if a valid month is selected
    if selected_month == "Velg måned":
        st.warning("Vennligst velg en måned for å fortsette.")
        return
    else:
        month_number = months.index(selected_month)
        month_days = hent_maned_dager(year, month_number)
    
        if not month_days:
            st.warning(f"Ingen produksjonsdager funnet for {selected_month} {year}.")
        else:
            st.write(f"Fant {len(month_days)} produksjonsdager for {selected_month} {year}")
    
    nedtrekk = ["Velg alternativ","Alle grafene","Kun månedlig gjennomsnitt"]
    graf_valg = st.selectbox("Vil du ha alle grafene eller kun månedlig gjennomsnitt?",nedtrekk)
    
    year = str(year)
    tittel = selected_month + " " + year
    
    daglig_data = []
    for dag in month_days:
        row = df[df['Dato'] == dag].iloc[0]
        
        # Print in Norwegian style
        stopptid = beregn_stopptid(row)
        arbeidstimer, antall_fisk = beregn_faktiskproduksjon(row)
        if stopptid is not None and arbeidstimer is not None and antall_fisk is not None:
            daglig_data.append((dag, stopptid, arbeidstimer, antall_fisk))
    
    if graf_valg == "Velg alternativ":
        st.warning("Vennligst velg et alternativ for å fortsette.")
        return
    
    elif graf_valg == "Alle grafene":
        
        for i in range(len(daglig_data)):
            # Plot daily graph
            st.write(f"Total stopptid i minutter: {round(daglig_data[i][1],2)}")
            st.write(f"Totale arbeidstimer: {round(daglig_data[i][2]/60,2)}")
            stopptid_impact = daglig_data[i][1] * oee_100
            stopptid_takt = round(stopptid_impact / daglig_data[i][2], 2)
            faktisk_takt = round(daglig_data[i][3] / daglig_data[i][2], 2)
            kjente_faktorer = round(stopptid_takt, 2)
            annet = oee_100 - kjente_faktorer - faktisk_takt
            annet = round(annet, 2)
            graf_type = "enkeltgraf"
            
            lag_graph(
                annet,
                faktisk_takt, stopptid_takt,
                daglig_data[i][0], graf_type
                )

    # Print separator for monthly average
    st.write("---")
    st.header(f"Oppsummering for {selected_month} {year}")

    # Calculate monthly averages
    avg_stopptid = np.mean([data[1] for data in daglig_data])
    avg_arbeidstimer = np.mean([data[2] for data in daglig_data])
    avg_antall_fisk = np.mean([data[3] for data in daglig_data])

    avg_stopptid_impact = avg_stopptid * oee_100
    avg_stopptid_takt = round(avg_stopptid_impact / avg_arbeidstimer, 2)
    avg_faktisk_takt = round(avg_antall_fisk / avg_arbeidstimer, 2)
    avg_kjente_faktorer = round(avg_stopptid_takt, 2)
    avg_annet = oee_100 - avg_kjente_faktorer - avg_faktisk_takt
    avg_annet = round(avg_annet, 2)
    graf_type = "manedsnitt"
    
    lag_graph(
        avg_annet,
        avg_faktisk_takt, avg_stopptid_takt,
        tittel, graf_type)
        
    
def main():
    st.title("Produksjonsanalyse")
    global sheet_type
    sheet_type = st.selectbox("Velg type ark:", ["slakt", "filet"])


    uploaded_file = st.file_uploader(f"Velg en Excel-fil (må være et 'input-{sheet_type}'-ark).", type=["xlsx"])
    analysis_type = st.selectbox("Velg analyse:", ["Spesifikk dato", "Ukesrapport", "Månedsrapport"])
    global oee_100
    oee_100 = 150 if sheet_type == "slakt" else 25
    global stiplet_hoyde
    stiplet_hoyde = 120 if sheet_type == "slakt" else 20

    if uploaded_file is None:
        st.warning("Vennligst last opp en Excel-fil for å fortsette.")
        return

    # Load and process the data
    global df
    df = les_data(uploaded_file)
    if df is None:
        st.warning("Ingen data tilgjengelig i den opplastede filen. Vennligst last opp en gyldig Excel-fil.")
        return

    # Ensure 'Dato' column exists
    try:
        df.iloc[:, 0] = pd.to_datetime(df.iloc[:, 0], format="%Y-%m-%d %H:%M:%S")
        df['Dato'] = df.iloc[:, 0].dt.date
    except ValueError as e:
        st.error(f"Feil ved behandling av datoer: {e}")
        return

    #------------------------------
    #          ENKELT DAG
    #------------------------------

    if analysis_type == "Spesifikk dato":
        
        enkelt_dato()
        
    #------------------------------
    #          UKESRAPPORT
    #------------------------------
    
    elif analysis_type == "Ukesrapport":
        
        uke()   
        
    #------------------------------    
    #         MÅNEDSRAPPORT
    #------------------------------
    else:            
        
        maned()
                
            
if __name__ == "__main__":
    main()






